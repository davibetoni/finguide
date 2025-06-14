import unicodedata
import os
from openpyxl import load_workbook
from datetime import datetime

MONTHS = {
    'Janeiro': 1, 'Fevereiro': 2, 'Março': 3, 'Abril': 4,
    'Maio': 5, 'Junho': 6, 'Julho': 7, 'Agosto': 8,
    'Setembro': 9, 'Outubro': 10, 'Novembro': 11, 'Dezembro': 12
}

CATEGORIES = {
    'Alimentação': ['PADARIA', 'RESTAURANTE', 'CONFEITARIA', 'PIZZARIA', 'LANCHES', 'CAFÉ', 'IFD', 'BEER', 'BAR', 'CHURRASCARIA', 'REST', 
             'FOOD', 'SANDUÍCHE', 'HAMBÚRGUER', 'PIZZA', 'SORVETE', 'MC DONALDS', 'BURGER KING', 'SUBWAY', 'FAST FOOD', 'PÃO DE AÇÚCAR'],
    'Saúde': ['HOSPITAL', 'CLÍNICA', 'MÉDICO', 'DENTISTA', 'SAÚDE', 'ALDEMIR', 'SAUDE'],
    'Educação': ['ESCOLA', 'UNIVERSIDADE', 'CURSO', 'AULA', 'LIVRARIA', 'PAPELARIA'],
    'Serviços': ['SERVIÇO', 'CONSULTORIA', 'ASSESSORIA', 'MANUTENÇÃO', 'REPARO', 'DUO'],
    'Utilidade': ['PRIME', 'NETFLIX', 'VIVO', 'YOUTUBE', 'GOOGLE', 'SPOTIFY'],
    'Transporte': ['UBER', '99', 'POSTO', 'PAREBEM', 'GASOLINA', 'ABASTECIMENTO', 'RODO'],
    'Veículo': ['CARRO', 'MOTO', 'VEÍCULO', 'AUTOPEÇAS', 'AUTO', 'DORO'],
    'Entreterimento': ['CINEMA', 'SHOW', 'TEATRO', 'EVENTO', 'SOCCER'],
    'Compras': ['LOJA', 'COMÉRCIO', 'VESTUÁRIO', 'CALÇADOS', 'ELETRÔNICOS', 'ALIEXPRESS', 'AMAZON', 'SHOPPING', 'ELETRO'],
    'Farmácia': ['DROGAL', 'FARMÁCIA'],
    'Pet': ['PET SHOP', 'ANIMAL', 'VETERINÁRIO', 'PET', 'AUFORNO'],
    'Mercado': ['MERCADO', 'SUPERMERCADO', 'LEVE PRA VIDA', 'ATACADÃO', 'CARREFOUR', 'BIG', 'PÃO DE AÇÚCAR'],
    'Casa': ['ALUGUEL', 'IPTU', 'CONDOMÍNIO', 'IMÓVEL', 'MORADIA', 'CASA'],
    'Outros': []
}



def get_data_file():
    return f"files/statement_summary_{datetime.now().year}.xlsx"

def months():
    return MONTHS

def list_months():
    return list(MONTHS.keys())

def get_month_number_by_name(month):
    return MONTHS.get(month, '01')

def get_month_name_by_number(month_number):
    for name, number in MONTHS.items():
        if number == month_number:
            return name
    return None

def list_categories():
    return list(CATEGORIES.keys())

def normalize_text(text):
    return unicodedata.normalize('NFKD', text).encode('ASCII', 'ignore').decode('utf-8')

def parse_amount(value_str):
    value_str = (
        value_str.replace('R$', '')
        .replace('.', '')
        .replace(',', '.')
        .strip()
        .rstrip('.')
    )
    return float(value_str)

def categorize(description):
    description = description.upper()
    for category, keywords in CATEGORIES.items():
        if any(keyword in description for keyword in keywords):
            return category
    return 'Outros'

def remove_sheet_if_exists(file_path, sheet_name="template"):
    if not os.path.exists(file_path):
        return

    wb = load_workbook(file_path)

    if sheet_name in wb.sheetnames:
        sheet_to_remove = wb[sheet_name]
        # Check how many visible sheets remain if we remove this one
        visible_sheets = [ws for ws in wb.worksheets if ws.sheet_state == 'visible' and ws.title != sheet_name]

        # If no other visible sheets remain, create a new one first
        if not visible_sheets:
            wb.create_sheet(title="Sheet1")

        wb.remove(sheet_to_remove)

    wb.save(file_path)